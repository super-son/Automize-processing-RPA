from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = Workbook() 
ws = wb.active
ws.title = "Student-Grade" 
ws.sheet_properties.tabColor = "ff33ff"

############################# 데이터 삽입

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
ws.append([1,10,8,5,14,26,12])
ws.append([2,7,3,7,15,24,18])
ws.append([3,9,5,8,8,12,4])
ws.append([4,7,8,7,17,21,18])
ws.append([5,7,8,7,16,25,15])
ws.append([6,3,5,8,8,17,0])
ws.append([7,4,9,10,16,27,18])
ws.append([8,6,6,6,15,19,17])
ws.append([9,10,10,9,19,30,19])
ws.append([10,9,8,8,20,25,20])

############################# 퀴즈 2의 점수를 모두 10으로 수정 #########################################3

# for row in ws.iter_rows(min_row=2) : # 조건컨트롤 가능. iter도 없어도 됨. 전체라는 뜻
#     row[3].value = 10

# 행 제목을 찾은 뒤 데이터 교체 방법
# for row in ws.iter_rows():
#     for cell in row:
#         if cell.value == "퀴즈2":
#             a = cell.column
# for row in ws.iter_rows() : 
#     row[a-1].value = 10

############################# H열에 총점 #####################################

# ws["H1"] = " 총점"

# for idx, row in enumerate(ws.iter_rows(min_row=2)) : # 조건컨트롤 가능. iter도 없어도 됨. 전체라는 뜻. 위에 제목인 초점을 먼저 정의해주고 만들어야되.
#     row[7].value = "=SUM(B{0}:G{1})".format(idx+2,idx+2)    

############################# I열에 성적정보 ##################################### 실패!!!!

# ws["I1"] = " 성적"

# for idx, row in enumerate(ws.iter_rows(min_row=2)) : # 조건컨트롤 가능. iter도 없어도 됨. 전체라는 뜻. 위에 제목인 초점을 먼저 정의해주고 만들어야되.
#     if row[7].value >= 90:
#         row[8].value = "A"
#     elif row[7].value >= 80:
#         row[8].value = "B"
#     elif row[7].value >= 70:
#         row[8].value = "C"
#     else:
#         row[8].value = "D"
############################################################## 나코딩이 하는 거.

for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목인 경우 skip
        continue
    cell.value = 10

ws["H1"] = " 총점"
ws["I1"] = " 성적"

scores=[
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20),
]

for idx, score in enumerate(scores, start=2): #idx를 2부터 시작하게 선언
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{0}:G{1})".format(idx,idx)
    #8=H
    # 애도 결국 총점 못구해서 score을 가공하네..
    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    if score[1] <5:
        grade = "F"
    ws.cell(row=idx, column=9).value = grade
    
############################ 이건 excel 배열과 같이 그대로 보여주는 코드. 컨트롤은 힘들지 #######################333
for row in ws.rows:
    for cell in row:
        # 각 셀에 대해서 가로세로 중앙으로 예쁘게 설정
        cell.alignment = Alignment(horizontal="center", vertical="center")

for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

wb.save("scores.xlsx")

