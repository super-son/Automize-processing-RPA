from openpyxl import load_workbook 
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# 보더와 사이드는 테두리 때메 # 패턴필은 글자색이 아닌 배경색!!! # 얼라이먼트는 글자위치변경
wb = load_workbook("sample.xlsx") 
ws = wb.active 

# 번호, 영어, 수학
a1  = ws["A1"]
b1  = ws["B1"]
c1  = ws["C1"]

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50
# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 빨갛게, 이탤릭, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 핑크색, Arial 폰트, 취소선
c1.font = Font(color="0000FF", size=20, underline="single") # 더블도 있어

# 테두리 적용
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90 점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 셀에 대해서 가로세로 중앙으로 예쁘게 설정
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # 센터도 되고 left, right, top, bottom 다 되

        if cell.column ==1: # A번호열은 제외
            continue
        # 개꿀팁 코드. 이것이 정수형이라면! 이라는 뜻
        if isinstance(cell.value, int) and cell.value > 90: 
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # '배경'을 초록색으로
            cell.font = Font(color="FF0000") # 폰트 색상 변경

# 엑셀에 보기 > 틀 고정이라는 것을 이용하면 첫 행은 스크롤해도 고정된다. 구현해보자
ws.freeze_panes = "B2" # B2기준으로 틀 고정. 걔를 기준으로 하면 오른쪽, 아래 스크롤 해보면 스무스하게 고정이되네!
wb.save("sample_style.xlsx")