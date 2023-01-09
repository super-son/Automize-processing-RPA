from openpyxl import Workbook
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx")
wb = Workbook() 
ws = wb.active
ws.title = "test"

# 데이터 열 삽입
ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
# 데이터 삽입
ws["I1"] = " 성적"

for row in ws.iter_rows(min_row=2,max_col=5) : # 조건컨트롤 가능. iter도 없어도 됨. 전체라는 뜻
    row[3].value = 10

# 행 제목을 찾은 뒤 그 줄들 데이터 교체 방법
for row in ws.iter_rows():
    for cell in row:
        if cell.value == "퀴즈2":
            a = cell.column
for row in ws.iter_rows() : 
    row[a-1].value = 10

# D열 콕 찝어서 그 줄들 데이터 교체 방법
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목인 경우 skip
        continue
    cell.value = 10

#####################################################################################

# 전체에 대한 설정
for row in ws.rows:
    for cell in row:
        # 각 셀에 대해서 가로세로 중앙으로 예쁘게 설정
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 터미널에서 엑셀 그대로 값들을 보여줌
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

wb.save("test.xlsx")
