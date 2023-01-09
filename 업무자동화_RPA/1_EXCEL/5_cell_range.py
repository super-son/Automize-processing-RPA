from openpyxl import Workbook # 파일 불러오기
from random import *

wb = Workbook()
ws = wb.active 

# 이젠 크게크게 자료를 넣어본다
# 1줄씩 데이터 넣기 # append로 row기준으로 이렇게 넣을수 있네! 
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])

# col_B = ws["B"] # 영어 column만 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
# for cols in col_range: # 튜플로 묶여있기 때문에 이중 for문 안쓰면 value값 못받아
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 1번째 row만 가지고 오기(title)
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기 (슬라이싱과 다른거야! 6포함한다)
# for rows in row_range:
#     for cell in rows: 
#         print(cell.value, end= " ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end= " ")
#         # print(cell.coordinate, end=" ") # 각 셀에 대한 좌표정보를 보여주는 코드
#         xy = coordinate_from_string(cell.coordinate) # 좌표가 보기 불편하다면 이 코드를 이용해
#         print(xy, end=" ") 
#         print(xy[0],end=" ")
#         print(xy[1],end=" ") # 이런식으로 세분화해서 컨트롤 가능하다는 점
#     print()

# 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[2].value)

# 전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows() : # 전체 row라는 뜻
#     print(row[1].value)
# for column in ws.iter_rows(): # 전체 column라는 뜻
#     print(column[0].value)

########################################################################################################

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지의 범위를 지정해주는 개꿀코드(B1-C5의 범위가 되겠네)
# 그리면서 생각을 해봐. 조금 헷갈리긴함. min-max 지정하지않으면 알아서 최소최대로 지정한다는것.
# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3) : 
#     print(row[1].value) # 이렇게하면 그치 (C1-C5가 되겠네)
# 응용
# for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3) : 
#     print(col[1].value) # 이렇게하면 그치 (A2-C2가 되겠네)

wb.save("sample.xlsx")