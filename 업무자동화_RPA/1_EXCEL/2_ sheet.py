from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff33ff" # RGB 형태로 값을 넣어주면 탭 색상 변경. 얘는 구글에 rgb 검색해서 w3c꺼 코드 복붙


ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성
new_ws = wb["NewSheet"] # dict 형태로 sheet에 접근이 가능
# 그 후 new_ws .....

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test" # a1 셀. 열은 A, B, C ... 행은 1, 2, 3 ...
targer = wb.copy_worksheet(new_ws) # 복사
targer.title = "Copied Sheet" # 네이밍

# 이게 마지막으로 쓰는 문법인듯
wb.save("sample.xlsx")