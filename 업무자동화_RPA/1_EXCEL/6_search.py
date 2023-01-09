from openpyxl import load_workbook 
wb = load_workbook("sample.xlsx") 
ws = wb.active 
##################################불러오기

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

## 데이터 수정코드.. 그리고 [0] 이런게 아니라면 튜플값에 대한 거니까 이중 for문이라고 생각하면 될듯.
# for row in ws.iter_rows(max_row=1):
#     for cell in row:
#         if cell.value == "영어":
#             cell.value = "컴퓨터"

# 위에랑 같은 뜻이지 ㅋ
for row in ws.iter_rows(max_row=1):
    row[1].value = "컴퓨터"

wb.save("sample_modified.xlsx")

