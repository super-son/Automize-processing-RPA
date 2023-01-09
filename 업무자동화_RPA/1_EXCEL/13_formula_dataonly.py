from openpyxl import load_workbook
# wb= load_workbook("sample_formula.xlsx") 
# ws = wb.active
# # 수식 그대로 가져오고 있음
# for row in ws.values: # 이렇게 적으면 각 셀의 객체가 아닌 value 정보를 바로 따오는..
#     for cell in row:
#         print(cell)

wb= load_workbook("sample_formula.xlsx",data_only=True) 
ws = wb.active
# 수식이 아닌 실제 데이터 값을 출력해봄
# evaluate 되지 않은 상태의 데이터는 None이거든
# 그러니까 우리가 코딩을 통해서 식을 만들어준거니까 evaluate되지않았는데 파일에 들어갔다 나가면 저장하겠냐고 물어본다. 
# 그때 저장을 해야 엑셀에 수식이 계산된 채로 있는것이다. 그 후 다시 print 하면 none이 안뜸

###### 주의할 점. openpyxl을 이용해서 수식을 이용할 때 그 파일은 사람이 한번 여는것을 전제로 해야지. 여기저기서 타고타는 방식은 에러가 날수있지.

for row in ws.values: # 이렇게 적으면 각 셀의 객체가 아닌 value 정보를 바로 따오는..
    for cell in row:
        print(cell)

