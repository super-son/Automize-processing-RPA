import datetime
from openpyxl import Workbook
wb= Workbook() # 얘는 sample.xlsx 가져온게 아니고 이렇게 새로만든거임
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["A2"] = "=SUM(1,2,3)" # 6이 나오게된다
ws["A3"] = "=AVERAGE(1,2,3)" # 2가 나오게 된다

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"


wb.save("sample_formula.xlsx")