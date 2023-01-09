from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# 해당 셀에 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을땐 none

# row = 1, 2, 3, ..
# column = A(1), B(2), C(3), ..
# ws.cell(row=1, column=1)는 A1과 같다.
print(ws.cell(row=1, column=1)) # = print(ws["A1"])
print(ws.cell(row=1, column=2).value) # = print(ws["B1"].value)

c = ws.cell(column=3, row=1, value=10) # = ws["C1"].value = 10
print(c.value) # 10이 나오지

from random import *
# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1,11): # 10개 row
    for y in range(1,11): # 10개 칼럼
        ws.cell(row=x, column=y, value=randint(0,100)) # 0부터 100까지의 랜덤숫자
        ws.cell(row=x, column=y, value=index) # index로 덮어쓰기 하기
        index += 1

wb.save("sample.xlsx") # 새로 덮어쓰기를 하는구만
