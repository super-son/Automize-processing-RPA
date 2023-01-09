from openpyxl import load_workbook 
wb = load_workbook("sample.xlsx") 
ws = wb.active 

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1) # 오른쪽으로 1칸씩만 땡기는 거지 rows=3 이라면 밑으로 3칸옮기기. 마이너스로 하면 왼쪽이나 위!!!!
# 기존의 데이터가 있다면 덮어쓰기로 들어가!
ws["B1"].value = "국어" # B!셀에 '국어'입력

wb.save("sample_korean.xlsx")